import requests
import urllib3
import time
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


green_color = "\033[92m"
red_color = "\033[91m"
reset_color = "\033[0m"



print(green_color + """
  _   _                           _____                       _            
 | \ | | ___  ___ ___ _   _ ___  | ____|_  ___ __   ___  _ __| |_ ___ _ __ 
 |  \| |/ _ \/ __/ __| | | / __| |  _| \ \/ / '_ \ / _ \| '__| __/ _ \ '__|
 | |\  |  __/\__ \__ \ |_| \__ \ | |___ >  <| |_) | (_) | |  | ||  __/ |   
 |_| \_|\___||___/___/\__,_|___/ |_____/_/\_\ .__/ \___/|_|   \__\___|_|   
                                            |_|
                                            """ + red_color + """- Created by HSKY                            
""" + reset_color)

base_url = "https://localhost:8834"
verify = False 
access_key = ""
secret_key = ""

def export_scan(scan_id):
    url = f"{base_url}/scans/{scan_id}/export"
    headers = {
        "X-ApiKeys": f"accessKey={access_key}; secretKey={secret_key}"
    }
    data = {
        "format": "nessus",
        "chapters": "vuln_hosts_summary"
    }
    response = requests.post(url, headers=headers, data=data, verify=verify)
    if response.status_code == 200:
        file_id = response.json().get("file")
        print(f"Exported scan successfully. File ID: {file_id}")
        return file_id
    else:
        print(f"Failed to export scan. Status code: {response.status_code}")
        print(response.text)  
        return None

def download_exported_scan_with_retry(scan_id, file_id, max_retries=10, retry_interval=10):
    retry_count = 0
    while retry_count < max_retries:
        url = f"{base_url}/scans/{scan_id}/export/{file_id}/download"
        headers = {
            "X-ApiKeys": f"accessKey={access_key}; secretKey={secret_key}"
        }
        response = requests.get(url, headers=headers, verify=verify)
        if response.status_code == 200:
            with open(f"exported_scan_{scan_id}.nessus", "wb") as f:
                f.write(response.content)
            print(f"Downloaded exported scan file for scan ID {scan_id}")
            return f"exported_scan_{scan_id}.nessus"
        elif response.status_code == 409:
            print("Report is still being generated. Retrying...")
            time.sleep(retry_interval)
            retry_count += 1
        else:
            print(f"Failed to download exported scan file for scan ID {scan_id}. Status code: {response.status_code}")
            print(response.text) 
            return None

    print(f"Maximum retry limit reached. Failed to download exported scan file for scan ID {scan_id}.")

def extract_vulnerability_info(xml_file, scan_id, scan_name):
    
    tree = ET.parse(xml_file)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active

    ws.append(["Plugin Name", "Risk Factor", "Solution", "Synopsis", "Affected Port", "Service Name", "Host IP"])

    column_titles = ["Plugin Name", "Risk Factor", "Solution", "Synopsis", "Affected Port", "Service Name", "Host IP"]
    ws.append(column_titles)
    for cell in ws[1]:
        cell.font = Font(size=16, bold=True, color="0000FF")

    grouped_vulnerabilities = {}

    for report_host in root.findall(".//ReportHost"):
        
        host_ip = report_host.attrib.get("name")

        for report_item in report_host.findall(".//ReportItem"):
            plugin_name = report_item.get("pluginName")
            risk_factor = report_item.find("risk_factor").text
            if risk_factor == "None":
                continue
            solution = report_item.find("solution").text
            synopsis = report_item.find("synopsis").text
            affected_port = report_item.get("port")
            service_name = report_item.get("svc_name", "")

            if affected_port == "0":
                continue

            key = (plugin_name, risk_factor, solution, synopsis)

            if key in grouped_vulnerabilities:
                existing_affected_ports = set(grouped_vulnerabilities[key][4])
                existing_service_names = set(grouped_vulnerabilities[key][5])
                existing_host_ips = set(grouped_vulnerabilities[key][6])

                if affected_port not in existing_affected_ports:
                    existing_affected_ports.add(affected_port)
                    grouped_vulnerabilities[key][4] = list(existing_affected_ports)
                
                if service_name and service_name not in existing_service_names:
                    existing_service_names.add(service_name)
                    grouped_vulnerabilities[key][5] = list(existing_service_names)
                
                if host_ip not in existing_host_ips:
                    existing_host_ips.add(host_ip)
                    grouped_vulnerabilities[key][6] = list(existing_host_ips)
            else:
                grouped_vulnerabilities[key] = [plugin_name, risk_factor, solution, synopsis, [affected_port], [service_name], [host_ip]]

    for key, value in grouped_vulnerabilities.items():
        plugin_name, risk_factor, solution, synopsis, affected_ports, service_names, host_ips = value
        affected_ports_str = ",".join(affected_ports) if affected_ports[0] != "0" else ""
        service_names_str = ",".join(service_names)
        host_ips_str = ",".join(host_ips)
        ws.append([plugin_name, risk_factor, solution, synopsis, affected_ports_str, service_names_str, host_ips_str])

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor="E6E6FA")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 2:
                if cell.value == "Critical":
                    cell.font = Font(color="FFFFFF") 
                    cell.fill = PatternFill(fill_type="solid", fgColor="6A329F")
                elif cell.value == "High":
                    cell.font = Font(color="FFFFFF")
                    cell.fill = PatternFill(fill_type="solid", fgColor="FF0000") 
                elif cell.value == "Medium":
                    cell.font = Font(color="000000")
                    cell.fill = PatternFill(fill_type="solid", fgColor="FF8000")
                elif cell.value == "Low":
                    cell.font = Font(color="000000")
                    cell.fill = PatternFill(fill_type="solid", fgColor="8FCE00")

    excel_file = f"{selected_scan_id}_{selected_scan_name.replace('/', '-')}.xlsx"
    wb.save(excel_file)
    print(f"Vulnerability information has been written to '{excel_file}'.")
    

def get_scan_ids_and_names():
    url = f"{base_url}/scans"
    headers = {
        "X-ApiKeys": f"accessKey={access_key}; secretKey={secret_key}"
    }
    response = requests.get(url, headers=headers, verify=verify)
    if response.status_code == 200:
        scans = response.json().get("scans", [])
        scan_ids_and_names = [(scan["id"], scan["name"]) for scan in scans]
        return scan_ids_and_names
    else:
        print(f"Failed to fetch scan IDs and names. Status code: {response.status_code}")
        print(response.text)
        return None

def select_scan_id(scan_ids_and_names):
    print("Available Scan IDs and Names:")
    for idx, (scan_id, scan_name) in enumerate(scan_ids_and_names, start=1):
        print(f"{idx}. {scan_id} - {scan_name}")
    selected_idx = int(input(red_color + "Enter the index of the scan you want to export and download: " + reset_color)) - 1
    return scan_ids_and_names[selected_idx]

scan_ids_and_names = get_scan_ids_and_names()
if scan_ids_and_names:
    selected_scan_id, selected_scan_name = select_scan_id(scan_ids_and_names)
    file_id = export_scan(selected_scan_id)
    if file_id:
        nessus_file_path = download_exported_scan_with_retry(selected_scan_id, file_id)
        if nessus_file_path:
            extract_vulnerability_info(nessus_file_path, selected_scan_id, selected_scan_name)
